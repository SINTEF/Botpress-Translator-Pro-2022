# Python Script to manage translation of botpress chatbots

# Arguments:
# 2 modes, either extract or pack
# Takes a botpress chatbot path argument (.tgz file)
# Extract mode: Generate an excel file with all the translations needed
# Pack mode: Generate a new botpress chatbot from an excel file and the original chatbot
# Source language: The language of the original chatbot, default to english
# Target language: The language of the new chatbot, default to french

import argparse
from distutils.fancy_getopt import wrap_text
import glob
import json
import tarfile
import tempfile
import openpyxl

from translate import translate

args = argparse.ArgumentParser()
args.add_argument(
    "-m",
    "--mode",
    help="Mode of operation",
    choices=["extract", "pack"],
    default="extract",
)
args.add_argument(
    "-b",
    "--bot",
    help="Path to the botpress chatbot to extract or pack",
    default="*.tgz",
)
args.add_argument(
    "-e",
    "--excel",
    help="Path to the excel file to extract or pack",
    default="translations.xlsx",
)
args.add_argument(
    "-s",
    "--source",
    help="Source language of the chatbot, default to english",
    default="en",
)
args.add_argument(
    "-t",
    "--target",
    help="Target language of the chatbot, default to french",
    default="fr",
)


def extract(bot_path, excel_path, source, target):
    # If the bot path contains an star, find the first file that matches the pattern
    if "*" in bot_path:
        bot_path = glob.glob(bot_path)[0]
        if not bot_path:
            print("No bot found for path " + bot_path)
            return

    print("Loading texts from bot " + bot_path)
    # Extract the bot to a temporary directory
    with tempfile.TemporaryDirectory() as temporary_directory:
        # Extract the .tgz file in python
        with tarfile.open(bot_path, "r:gz") as tar:
            tar.extractall(temporary_directory)
        # Open content-elements/builtin_text.json
        with open(
            temporary_directory + "/content-elements/builtin_text.json", "r"
        ) as f:
            # read and parse json
            builtin_texts = json.load(f)
            # builtin_text is an array of objects containing an id and a formData object
            entries = [
                (entry["id"], entry["formData"]["text$en"]) for entry in builtin_texts
            ]

        # Open content-elements/builtin_card.json
        with open(
            temporary_directory + "/content-elements/builtin_card.json", "r"
        ) as f:
            # read and parse json
            builtin_cards = json.load(f)
            # builtin_card is an array of objects containing an id and a formData object
            for builtin_card in builtin_cards:
                entries.append(
                    (
                        builtin_card["id"] + ".title",
                        builtin_card["formData"]["title$en"],
                    )
                )
                if "subtitle$en" in builtin_card["formData"]:
                    entries.append(
                        (
                            builtin_card["id"] + ".subtitle",
                            builtin_card["formData"]["subtitle$en"],
                        )
                    )
                for index, option in enumerate(builtin_card["formData"]["actions$en"]):
                    entries.append(
                        (
                            builtin_card["id"] + ".actionTitle[" + str(index) + "]",
                            option["title"],
                        )
                    )
        # Open content-elements/builtin_carousel.json
        with open(
            temporary_directory + "/content-elements/builtin_carousel.json", "r"
        ) as f:
            # read and parse json
            builtin_carousels = json.load(f)
            # builtin_carousel is an array of objects containing an id and a formData object
            for builtin_carousel in builtin_carousels:
                for index, item in enumerate(builtin_carousel["formData"]["items$en"]):
                    entries.append(
                        (
                            builtin_carousel["id"] + ".itemTitle[" + str(index) + "]",
                            item["title"],
                        )
                    )
                    for actionIndex, option in enumerate(item["actions"]):
                        entries.append(
                            (
                                builtin_carousel["id"]
                                + ".itemActionTitle["
                                + str(index)
                                + "]["
                                + str(actionIndex)
                                + "]",
                                option["title"],
                            )
                        )
        # Open content-elements/builtin_image.json
        with open(
            temporary_directory + "/content-elements/builtin_image.json", "r"
        ) as f:
            # read and parse json
            builtin_images = json.load(f)
            # builtin_image is an array of objects containing an id and a formData object
            for builtin_image in builtin_images:
                if "title$en" in builtin_image["formData"]:
                    entries.append(
                        (
                            builtin_image["id"] + ".title",
                            builtin_image["formData"]["title$en"],
                        )
                    )

        # Open content-elements/builtin_single-choice.json
        with open(
            temporary_directory + "/content-elements/builtin_single-choice.json", "r"
        ) as f:
            # read and parse json
            builtin_single_choices = json.load(f)
            # builtin_single-choice is an array of objects containing an id and a formData object
            for builtin_single_choice in builtin_single_choices:
                entries.append(
                    (
                        builtin_single_choice["id"] + ".dropdown",
                        builtin_single_choice["formData"]["dropdownPlaceholder$en"],
                    )
                )
                if "title$en" in builtin_single_choice["formData"]:
                    entries.append(
                        (
                            builtin_single_choice["id"] + ".title",
                            builtin_single_choice["formData"]["title$en"],
                        )
                    )
                for index, option in enumerate(
                    builtin_single_choice["formData"]["choices$en"]
                ):
                    entries.append(
                        (
                            builtin_single_choice["id"] + ".choice[" + str(index) + "]",
                            option["title"],
                        )
                    )
        # We don't translate the videos automatically because they most likely need to be changed

        # Open content-elements/dropdown.json
        with open(temporary_directory + "/content-elements/dropdown.json", "r") as f:
            # read and parse json
            dropdowns = json.load(f)
            # dropdown is an array of objects containing an id and a formData object
            for dropdown in dropdowns:
                entries.append(
                    (
                        dropdown["id"] + ".message",
                        dropdown["formData"]["message$en"],
                    )
                )
                entries.append(
                    (
                        dropdown["id"] + ".placeholderText",
                        dropdown["formData"]["placeholderText$en"],
                    )
                )
                for index, option in enumerate(dropdown["formData"]["options$en"]):
                    entries.append(
                        (
                            dropdown["id"] + ".option[" + str(index) + "]",
                            option["label"],
                        )
                    )

        # Translate the text using Google Translate API
        print("Translating texts...")
        texts_to_translate = [text for _, text in entries]
        translated_texts = translate(texts_to_translate, source, target)

        print("Writing Excel file...")
        # Create a new excel file
        wb = openpyxl.Workbook()
        # Create a new sheet
        ws = wb.active
        ws.title = "Translations"
        # Write the entries to the sheet
        ws.append(["Identifier", "Original English Text", "Translation"])
        # ws.append(entries)
        for (id, text) in entries:
            ws.append([id, text, translated_texts[text]])

        idStyle = openpyxl.styles.NamedStyle(name="id")
        idStyle.font = openpyxl.styles.Font(
            bold=True,
            name="Courier New",
            size=9,
        )
        idStyle.fill = openpyxl.styles.PatternFill(
            fgColor="EEEEEE",
            fill_type="solid",
        )
        idStyle.alignment = openpyxl.styles.Alignment(
            wrap_text=True,
        )
        headerStyle = openpyxl.styles.NamedStyle(name="header")
        headerStyle.font = openpyxl.styles.Font(
            bold=True,
            size=10,
            name="Calibri",
        )
        headerStyle.fill = openpyxl.styles.PatternFill(
            fgColor="CCCCCC",
            fill_type="solid",
        )
        originalStyle = openpyxl.styles.NamedStyle(name="original")
        originalStyle.font = openpyxl.styles.Font(
            name="Calibri",
            # Dark brown colour
            color="663300",
            italic=True,
        )
        originalStyle.alignment = openpyxl.styles.Alignment(
            vertical="top",
            wrap_text=True,
        )
        translationStyle = openpyxl.styles.NamedStyle(name="translation")
        translationStyle.font = openpyxl.styles.Font(
            name="Calibri",
        )
        translationStyle.alignment = openpyxl.styles.Alignment(
            vertical="top",
            wrap_text=True,
        )
        ws["A1"].style = idStyle
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30

        # Lock the cells that shouldn't be edited
        ws.protection.sheet = True

        # Set headers styles
        for row in ws["A1":"C1"]:
            for cell in row:
                cell.style = headerStyle
        # Set identifier styles
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.style = idStyle
        # Set original text style
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.style = originalStyle
        # Set translation styles
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.style = translationStyle
                # Unlock the translation rows
                cell.protection = openpyxl.styles.protection.Protection(locked=False)
        wb.save(excel_path)
        print("Excel file created: " + excel_path)
        print("Done 🥳")

if __name__ == "__main__":
    args = args.parse_args()
    if args.mode == "extract":
        extract(args.bot, args.excel, args.source, args.target)
    elif args.mode == "pack":
        pack(args.bot, args.excel)
